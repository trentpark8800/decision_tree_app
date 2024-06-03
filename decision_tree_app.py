from typing import List, TypeVar, Dict
from io import BytesIO

import streamlit as st
from pandas.core.frame import DataFrame
from matplotlib.figure import Figure

from openpyxl import load_workbook

from data_management.storage import S3StorageManager

from data_modelling.utils import (
    create_column_encoders,
    encode_columns,
    read_xl_data_into_dataframe,
    get_encoder_mapping,
)

from data_modelling.decision_tree import (
    plot_decision_tree,
    evaluate_decision_tree,
    create_decision_tree,
)


def file_change():
    st.session_state["file_change"] = True
    st.session_state["data_change"] = True


def store_uploaded_data_in_blob(data: st.file_uploader) -> str:
    """_summary_

    Returns:
        [BlobStorageManager, str]: _description_
    """
    storage_manager: S3StorageManager = st.session_state["storage_manager"]
    blob_name: str = storage_manager.upload_data_to_s3(data)
    st.session_state["file_change"] = False
    return blob_name


def get_excel_sheet_names() -> List[str]:
    """ """
    file_like_object: BytesIO = BytesIO(
        st.session_state["storage_manager"].retrieve_file_from_s3(
            st.session_state["blob_file_name"]
        )
    )
    wb = load_workbook(file_like_object, read_only=True)

    sheet_names = wb.sheetnames.copy()

    del file_like_object

    return sheet_names


def load_blob_data_into_dataframe(sheet_name: str) -> DataFrame:
    df = read_xl_data_into_dataframe(
        st.session_state["storage_manager"].retrieve_file_from_s3(
            st.session_state["blob_file_name"]
        ),
        sheet_name=sheet_name,
    )

    st.session_state["data_change"] = False

    return df


def main():
    """
    Entry point for the script.
    """
    if "storage_manager" not in st.session_state:
        st.session_state["storage_manager"] = S3StorageManager()

    if "file_change" not in st.session_state:
        st.session_state["file_change"] = False

    if "data_change" not in st.session_state:
        st.session_state["data_change"] = False

    if "raw_train_df" not in st.session_state:
        st.session_state["raw_train_df"] = None

    if "raw_test_df" not in st.session_state:
        st.session_state["raw_test_df"] = None

    if "excel_sheets_names" not in st.session_state:
        st.session_state["excel_sheets_names"] = None

    if "testing_sheet_selection" not in st.session_state:
        st.session_state["testing_sheet_selection"] = None

    st.title("Decision Tree Model Simulator")
    st.markdown(
        """
        A simple demo to illustrate how decision trees work.
        """
    )

    # TODO: Remove this test code
    # st.write(st.session_state)

    data: st.file_uploader = st.file_uploader(
        label="Drop your .xlsx file here", key="data_upload", on_change=file_change
    )

    if st.session_state["file_change"] and st.session_state["data_upload"]:
        st.session_state["blob_file_name"] = store_uploaded_data_in_blob(data)

    if bool(st.session_state.data_upload):

        if st.session_state["data_change"]:
            st.session_state["excel_sheet_names"] = get_excel_sheet_names()

        training_sheet: str = st.selectbox(
            label="Choose the sheet to use for training",
            options=(st.session_state["excel_sheet_names"]),
            key="sheet_selection",
        )

        if bool(st.session_state["sheet_selection"]):
            st.session_state["raw_train_df"] = load_blob_data_into_dataframe(
                sheet_name=training_sheet
            )

        raw_train_df = st.session_state["raw_train_df"]

        st.header("Your Data")
        st.dataframe(raw_train_df.astype(str))

        # Select target variable
        target: str = st.selectbox(
            label="Choose your target variable",
            options=(raw_train_df.columns),
            key="target_var",
        )

        # Select features to be used
        features: List[str] = list(
            st.multiselect(
                label="Choose your features", options=(raw_train_df.columns), key="feature_var"
            )
        )

        selected_columns: List[str] = features.copy()
        selected_columns.append(target)

        # Encode string/object features and target var
        encoders: Dict = create_column_encoders(raw_train_df[selected_columns])
        train_df: DataFrame = encode_columns(raw_train_df[selected_columns], encoders)

        max_tree_depth: int = int(
            st.text_input(
                label="Specify the maximum depth of the tree",
                value="2",
                key="tree_depth_var",
            ),
        )

        info_purity_measure: str = st.selectbox(
            label="Specify what technique to use for information entropy",
            options=("entropy", "gini"),
            key="info_purity_var",
        )

        if st.session_state["target_var"] and st.session_state["feature_var"]:

            classifier = create_decision_tree(
                max_depth=max_tree_depth,
                info_purity_measure=info_purity_measure,
                train_df=train_df,
                features=features,
                target=target,
            )

            # Get class names so that the plot is labelled intuitively
            # If no encoding is required then simply use the unique labels as class names
            try:
                class_names: List[str] = list(
                    encoders[target].inverse_transform(sorted(train_df[target].unique()))
                )
            except KeyError:
                class_names: List[str] = list(sorted(train_df[target].astype("string").unique()))

            if st.checkbox(label="Create Decision Tree", key="create_decision_tree"):

                st.write("Tree created!")

                # Plot the decision tree on the streamlit app
                decision_tree_figure: Figure = plot_decision_tree(
                    classifier=classifier, features=features, class_names=class_names
                )
                st.pyplot(decision_tree_figure)

                # Let user view the encoder mappings
                encoder: str = st.selectbox(
                    label="Select encoded feature to view mapping between original label and encoded value",
                    options=(encoders.keys()),
                    key="encoder_mapping_option",
                )

                if encoder:
                    encoder_mapping_df: DataFrame = get_encoder_mapping(encoder, encoders)
                    st.dataframe(encoder_mapping_df.astype(str))

            if st.checkbox(label="Evaluate Training Metrics", key="evaluate_training_metrics"):
                confusion_matrix_fig, evaluation_metrics = evaluate_decision_tree(
                    classifier=classifier,
                    input_df=train_df,
                    features=features,
                    target=target,
                    class_names=class_names,
                )

                st.write(f"Decision tree accuracy: {evaluation_metrics.accuracy}")
                st.write(f"Decision tree precision: {evaluation_metrics.precision}")
                st.write(f"Decision tree F1 score: {evaluation_metrics.f1_score}")

                st.pyplot(confusion_matrix_fig)

            if bool(st.session_state["evaluate_training_metrics"]):
                testing_sheet: str = st.selectbox(
                    label="Choose the sheet to use for testing",
                    options=(st.session_state["excel_sheet_names"]),
                    key="testing_sheet_selection",
                )

            if bool(st.session_state["testing_sheet_selection"]):
                st.session_state["raw_test_df"] = load_blob_data_into_dataframe(
                    sheet_name=testing_sheet
                )
                test_df = encode_columns(
                    st.session_state["raw_test_df"][selected_columns], encoders
                )
                test_confusion_matrix_fig, test_evaluation_metrics = evaluate_decision_tree(
                    classifier=classifier,
                    input_df=test_df,
                    features=features,
                    target=target,
                    class_names=class_names,
                )

                st.write(f"Decision tree accuracy: {test_evaluation_metrics.accuracy}")
                st.write(f"Decision tree precision: {test_evaluation_metrics.precision}")
                st.write(f"Decision tree F1 score: {test_evaluation_metrics.f1_score}")

                st.pyplot(test_confusion_matrix_fig)


if __name__ == "__main__":
    main()
